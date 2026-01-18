"""Reports API routes - xlsx download."""

from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Assignment, Cage, Professor, Rack

router = APIRouter(prefix="/reports", tags=["reports"])

COST_PER_CAGE_DAY = 800


def create_xlsx_report(
    db: Session,
    start_date: date,
    end_date: date,
) -> BytesIO:
    """Create xlsx report with summary and detail sheets."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    number_format = "#,##0"
    currency_format = "#,##0원"

    # Get assignments in the date range
    assignments = (
        db.query(Assignment)
        .filter(
            Assignment.assigned_date >= start_date,
            Assignment.assigned_date <= end_date,
        )
        .order_by(Assignment.assigned_date, Assignment.cage_id)
        .all()
    )

    # Build lookups
    professors = {p.id: p for p in db.query(Professor).all()}
    cages = {c.id: c for c in db.query(Cage).all()}
    racks = {r.id: r for r in db.query(Rack).all()}

    # ========== [요약] Sheet ==========
    ws_summary = wb.active
    ws_summary.title = "요약"

    # Header
    summary_headers = ["교수명", "담당 학생", "사용 케이지 수", "총 비용"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Calculate professor summaries
    professor_summary: dict[int, int] = {}
    for assignment in assignments:
        professor_summary[assignment.professor_id] = (
            professor_summary.get(assignment.professor_id, 0) + 1
        )

    # Data rows
    row = 2
    total_cage_count = 0
    total_cost = 0

    for professor_id, cage_count in sorted(
        professor_summary.items(),
        key=lambda x: -x[1],
    ):
        professor = professors.get(professor_id)
        if not professor:
            continue

        cost = cage_count * COST_PER_CAGE_DAY
        total_cage_count += cage_count
        total_cost += cost

        ws_summary.cell(row=row, column=1, value=professor.name).border = thin_border
        ws_summary.cell(row=row, column=2, value=professor.student_name or "-").border = thin_border
        cell_count = ws_summary.cell(row=row, column=3, value=cage_count)
        cell_count.border = thin_border
        cell_count.number_format = number_format
        cell_count.alignment = Alignment(horizontal="right")
        cell_cost = ws_summary.cell(row=row, column=4, value=cost)
        cell_cost.border = thin_border
        cell_cost.number_format = currency_format
        cell_cost.alignment = Alignment(horizontal="right")
        row += 1

    # Total row
    total_font = Font(bold=True)
    ws_summary.cell(row=row, column=1, value="합계").font = total_font
    ws_summary.cell(row=row, column=1).border = thin_border
    ws_summary.cell(row=row, column=2, value="").border = thin_border
    cell_total_count = ws_summary.cell(row=row, column=3, value=total_cage_count)
    cell_total_count.font = total_font
    cell_total_count.border = thin_border
    cell_total_count.number_format = number_format
    cell_total_count.alignment = Alignment(horizontal="right")
    cell_total_cost = ws_summary.cell(row=row, column=4, value=total_cost)
    cell_total_cost.font = total_font
    cell_total_cost.border = thin_border
    cell_total_cost.number_format = currency_format
    cell_total_cost.alignment = Alignment(horizontal="right")

    # Column widths
    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 18
    ws_summary.column_dimensions["D"].width = 15

    # ========== [상세] Sheet ==========
    ws_detail = wb.create_sheet(title="상세")

    # Header
    detail_headers = ["날짜", "랙", "케이지 위치", "교수명", "담당 학생", "비용"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row = 2
    for assignment in assignments:
        professor = professors.get(assignment.professor_id)
        cage = cages.get(assignment.cage_id)
        rack = racks.get(cage.rack_id) if cage else None

        ws_detail.cell(row=row, column=1, value=assignment.assigned_date.strftime("%Y-%m-%d")).border = thin_border
        ws_detail.cell(row=row, column=2, value=rack.name if rack else "-").border = thin_border
        ws_detail.cell(row=row, column=3, value=cage.position if cage else "-").border = thin_border
        ws_detail.cell(row=row, column=4, value=professor.name if professor else "-").border = thin_border
        ws_detail.cell(row=row, column=5, value=professor.student_name if professor and professor.student_name else "-").border = thin_border
        cell_cost = ws_detail.cell(row=row, column=6, value=COST_PER_CAGE_DAY)
        cell_cost.border = thin_border
        cell_cost.number_format = currency_format
        cell_cost.alignment = Alignment(horizontal="right")
        row += 1

    # Column widths
    ws_detail.column_dimensions["A"].width = 12
    ws_detail.column_dimensions["B"].width = 12
    ws_detail.column_dimensions["C"].width = 12
    ws_detail.column_dimensions["D"].width = 15
    ws_detail.column_dimensions["E"].width = 15
    ws_detail.column_dimensions["F"].width = 12

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/download")
def download_report(
    start: date = Query(..., description="시작 날짜 (YYYY-MM-DD)"),
    end: date = Query(..., description="종료 날짜 (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    """Download xlsx report for the specified date range."""
    from urllib.parse import quote

    xlsx_file = create_xlsx_report(db, start, end)

    filename = f"케이지_사용_리포트_{start}_{end}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        xlsx_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )
